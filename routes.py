from flask import render_template, request, redirect, url_for, flash, session, abort, make_response, send_file, send_from_directory
from werkzeug.security import generate_password_hash
from flask_login import current_user
from werkzeug.utils import secure_filename
from sqlalchemy import extract, and_
from app import app, db
from models import User, Ticket, TicketComment, Attachment, MasterDataCategory, MasterDataPriority, MasterDataStatus, EmailSettings, TimezoneSettings, BackupSettings, EmailNotificationLog
from forms import LoginForm, TicketForm, UpdateTicketForm, CommentForm, UserRegistrationForm, AssignTicketForm, UserProfileForm, MasterDataCategoryForm, MasterDataPriorityForm, MasterDataStatusForm, EmailSettingsForm, TimezoneSettingsForm, BackupSettingsForm
from datetime import datetime
from utils.email import send_assignment_email  # Add this import
from utils.timezone import utc_to_ist
import logging
import os
import socket
import platform
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'csv', 'ppt', 'pptx'}
UPLOAD_FOLDER = 'uploads/'  # Set a secure uploads folder

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Helper function to check if user is logged in
def is_logged_in():
    return 'user_id' in session

# Helper function to get current user
def get_current_user():
    if is_logged_in():
        return User.query.get(session['user_id'])
    return None

# Helper function to require login
def login_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('common_login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Helper function to require super admin
def super_admin_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('common_login'))
        user = get_current_user()
        if not user or not user.is_super_admin:
            flash('Super Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Helper function to require HOD access
def hod_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('common_login'))
        user = get_current_user()
        if not user or not user.is_hod:
            flash('HOD access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Helper function to require admin or super admin access
def admin_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('common_login'))
        user = get_current_user()
        if not user or not user.can_manage_tickets:
            flash('Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def common_login():
    """Common login page for all user types"""
    if is_logged_in():
        user = get_current_user()
        # Redirect to appropriate dashboard based on role
        if user.is_super_admin:
            return redirect(url_for('super_admin_dashboard'))
        elif user.is_hod:
            return redirect(url_for('hod_dashboard'))
        elif user.is_admin:
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('user_dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            # Set session variables
            session['user_id'] = user.id
            session['role'] = user.role
            
            # Update IP address and system info
            user.ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR'))
            try:
                user.system_name = socket.gethostname()
            except:
                user.system_name = 'Unknown'
            db.session.commit()
            
            flash(f'Welcome back, {user.first_name}!', 'success')
            
            # Route to appropriate dashboard based on role
            if user.is_super_admin:
                return redirect(url_for('super_admin_dashboard'))
            elif user.is_hod:
                return redirect(url_for('hod_dashboard'))
            elif user.is_admin:
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('user_dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('common_login.html', form=form)

@app.route('/user-login', methods=['GET', 'POST'])
def user_login():
    """User login page - redirects to common login"""
    return redirect(url_for('common_login'))

@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page - redirects to common login"""
    return redirect(url_for('common_login'))

@app.route('/logout')
def logout():
    """Logout user"""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/user-dashboard')
@login_required
def user_dashboard():
    """User dashboard showing their tickets"""
    user = get_current_user()
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    search_query = request.args.get('search', '')
    
    # Build query
    query = Ticket.query.filter_by(user_id=user.id)
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    
    return render_template('user_dashboard.html', user=user, tickets=tickets, 
                         status_filter=status_filter, search_query=search_query)

@app.route('/user-profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    """User profile management - Super Admin can edit, users can only view"""
    user = get_current_user()
    form = UserProfileForm(obj=user)
    
    # Only allow Super Admins to edit profiles
    if request.method == 'POST' and not user.is_super_admin:
        flash('You do not have permission to edit profile information.', 'error')
        return redirect(url_for('user_profile'))
    
    if form.validate_on_submit() and user.is_super_admin:
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.email = form.email.data
        user.department = form.department.data
        user.specialization = form.specialization.data if form.specialization.data else None
        user.system_name = form.system_name.data
        
        # Handle profile image upload
        if 'profile_image' in request.files:
            file = request.files['profile_image']
            if file.filename != '':
                filename = secure_filename(file.filename)
                # Save file logic would go here
                user.profile_image = filename
        
        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('user_profile'))
    
    # Get user statistics for the profile
    user_tickets = Ticket.query.filter_by(user_id=user.id).all()
    total_tickets = len(user_tickets)
    open_tickets = len([t for t in user_tickets if t.status == 'Open'])
    resolved_tickets = len([t for t in user_tickets if t.status == 'Resolved'])
    
    user_stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'resolved_tickets': resolved_tickets,
        'recent_tickets': user_tickets[:5] if user_tickets else []
    }
    
    return render_template('user_profile.html', form=form, user=user, user_stats=user_stats, current_user=user)

@app.route('/hod-dashboard')
@hod_required
def hod_dashboard():
    """HOD dashboard showing department tickets"""
    user = get_current_user()
    if not user.is_hod:
        flash('HOD access required.', 'error')
        return redirect(url_for('index'))

    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    priority_filter = request.args.get('priority', 'all')
    category_filter = request.args.get('category', 'all')
    search_query = request.args.get('search', '')

    # Build query for department tickets - HOD can see all tickets from their department
    query = Ticket.query.join(User, Ticket.user_id == User.id).filter(User.department == user.department)

    if status_filter != 'all':
        query = query.filter(Ticket.status == status_filter)
    
    if priority_filter != 'all':
        query = query.filter(Ticket.priority == priority_filter)
    
    if category_filter != 'all':
        query = query.filter(Ticket.category == category_filter)
    
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))

    tickets = query.order_by(Ticket.created_at.desc()).all()

    # Department statistics
    dept_total_tickets = Ticket.query.join(User, Ticket.user_id == User.id).filter(User.department == user.department).count()
    dept_open_tickets = Ticket.query.join(User, Ticket.user_id == User.id).filter(User.department == user.department, Ticket.status == 'Open').count()
    dept_in_progress_tickets = Ticket.query.join(User, Ticket.user_id == User.id).filter(User.department == user.department, Ticket.status == 'In Progress').count()
    dept_resolved_tickets = Ticket.query.join(User, Ticket.user_id == User.id).filter(User.department == user.department, Ticket.status == 'Resolved').count()

    stats = {
        'dept_total_tickets': dept_total_tickets,
        'dept_open_tickets': dept_open_tickets,
        'dept_in_progress_tickets': dept_in_progress_tickets,
        'dept_resolved_tickets': dept_resolved_tickets
    }

    return render_template('hod_dashboard.html', 
                         user=user, 
                         tickets=tickets, 
                         stats=stats,
                         status_filter=status_filter,
                         priority_filter=priority_filter,
                         category_filter=category_filter,
                         search_query=search_query)

@app.route('/admin-dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard for ticket management"""
    user = get_current_user()
    if not user.can_manage_tickets:
        flash('Admin access required.', 'error')
        return redirect(url_for('index'))

    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    priority_filter = request.args.get('priority', 'all')
    category_filter = request.args.get('category', 'all')
    search_query = request.args.get('search', '')

    # Build query for all tickets (admins can see all tickets)
    query = Ticket.query

    if status_filter != 'all':
        query = query.filter(Ticket.status == status_filter)
    
    if priority_filter != 'all':
        query = query.filter(Ticket.priority == priority_filter)
    
    if category_filter != 'all':
        query = query.filter(Ticket.category == category_filter)
    
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))

    tickets = query.order_by(Ticket.created_at.desc()).all()

    # Admin statistics
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='Open').count()
    in_progress_tickets = Ticket.query.filter_by(status='In Progress').count()
    resolved_tickets = Ticket.query.filter_by(status='Resolved').count()
    assigned_to_me = Ticket.query.filter_by(assigned_to=user.id).count()

    stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'assigned_to_me': assigned_to_me
    }

    return render_template('admin_dashboard.html', 
                         user=user, 
                         tickets=tickets, 
                         stats=stats,
                         status_filter=status_filter,
                         priority_filter=priority_filter,
                         category_filter=category_filter,
                         search_query=search_query)

@app.route('/super-admin-dashboard')
@super_admin_required
def super_admin_dashboard():
    """Super Admin dashboard with full system overview and filters"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))

    # Comprehensive statistics (unfiltered)
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='Open').count()
    in_progress_tickets = Ticket.query.filter_by(status='In Progress').count()
    resolved_tickets = Ticket.query.filter_by(status='Resolved').count()
    total_users = User.query.filter_by(role='user').count()
    total_admins = User.query.filter_by(role='admin').count()
    hardware_tickets = Ticket.query.filter_by(category='Hardware').count()
    software_tickets = Ticket.query.filter_by(category='Software').count()

    stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'total_users': total_users,
        'total_admins': total_admins,
        'hardware_tickets': hardware_tickets,
        'software_tickets': software_tickets
    }

    # Filter parameters for recent tickets
    status_filter = request.args.get('status', 'all')
    priority_filter = request.args.get('priority', 'all')
    category_filter = request.args.get('category', 'all')
    search_query = request.args.get('search', '')
    day_filter = request.args.get('day', '')
    month_filter = request.args.get('month', '')
    year_filter = request.args.get('year', '')

    # Build filtered query for recent tickets
    query = Ticket.query
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    if priority_filter != 'all':
        query = query.filter_by(priority=priority_filter)
    if category_filter != 'all':
        query = query.filter_by(category=category_filter)
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))
    if year_filter:
        query = query.filter(extract('year', Ticket.created_at) == int(year_filter))
    if month_filter:
        query = query.filter(extract('month', Ticket.created_at) == int(month_filter))
    if day_filter:
        query = query.filter(extract('day', Ticket.created_at) == int(day_filter))

    recent_tickets = query.order_by(Ticket.created_at.desc()).limit(10).all()

    return render_template(
        'super_admin_dashboard.html',
        stats=stats,
        recent_tickets=recent_tickets,
        status_filter=status_filter,
        priority_filter=priority_filter,
        category_filter=category_filter,
        search_query=search_query,
        day_filter=day_filter,
        month_filter=month_filter,
        year_filter=year_filter
    )




@app.route('/create-ticket', methods=['GET', 'POST'])
@login_required
def create_ticket():
    """Create a new ticket"""
    form = TicketForm()
    user = get_current_user()

    if form.validate_on_submit():
        # Capture current session IP and system info for this specific ticket
        current_ip = request.environ.get('HTTP_X_FORWARDED_FOR',
                                         request.environ.get('HTTP_X_REAL_IP',
                                                             request.environ.get('REMOTE_ADDR')))

        # Get system name
        current_system_name = form.system_name.data.strip() if form.system_name.data and form.system_name.data.strip() else None
        if not current_system_name:
            if user.system_name and user.system_name.strip():
                current_system_name = user.system_name.strip()
            else:
                user_agent = request.headers.get('User-Agent', '').lower()
                timestamp = datetime.now().strftime('%m%d%H%M')
                
                if 'windows nt 10.0' in user_agent:
                    current_system_name = f'WIN10-PC-{timestamp}'
                elif 'windows nt 6.3' in user_agent:
                    current_system_name = f'WIN8-PC-{timestamp}'
                elif 'windows nt 6.1' in user_agent:
                    current_system_name = f'WIN7-PC-{timestamp}'
                elif 'windows' in user_agent:
                    current_system_name = f'WINDOWS-PC-{timestamp}'
                elif 'mac os x' in user_agent or 'macos' in user_agent:
                    current_system_name = f'MACOS-{timestamp}'
                elif 'linux' in user_agent and 'android' not in user_agent:
                    current_system_name = f'LINUX-{timestamp}'
                elif 'android' in user_agent:
                    current_system_name = f'ANDROID-{timestamp}'
                elif 'iphone' in user_agent:
                    current_system_name = f'IPHONE-{timestamp}'
                elif 'ipad' in user_agent:
                    current_system_name = f'IPAD-{timestamp}'
                else:
                    current_system_name = f'DEVICE-{timestamp}'

        user.ip_address = current_ip
        user.system_name = current_system_name

        # Handle file uploads (supporting multiple attachments)
        attachment_filenames = []
        files = request.files.getlist('image')  # Changed from 'attachments' to 'image'
        for file in files:
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                unique_filename = timestamp + filename
                os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
                try:
                    file.save(filepath)
                    attachment_filenames.append(unique_filename)
                    logging.info(f'Successfully uploaded file: {unique_filename}')
                except Exception as e:
                    flash(f'Error uploading file {filename}: {str(e)}', 'warning')
                    logging.error(f'Error uploading file {filename}: {str(e)}')

        # For backward compatibility, store the first image filename in image_filename, others in attachments
        image_filename = None
        other_attachments = []
        for fname in attachment_filenames:
            if any(fname.lower().endswith(ext) for ext in
                   ['.png', '.jpg', '.jpeg', '.gif', '.bmp']) and not image_filename:
                image_filename = fname
            else:
                other_attachments.append(fname)

        # Generate ticket number
        last_ticket = Ticket.query.order_by(Ticket.id.desc()).first()
        next_ticket_id = (last_ticket.id + 1) if last_ticket else 1
        ticket_number = f"GTN-{next_ticket_id:06d}"
        
        # Create the ticket first
        ticket = Ticket(
            ticket_number=ticket_number,
            title=form.title.data,
            description=form.description.data,
            category=form.category.data,
            priority=form.priority.data,
            user_id=user.id,
            user_name=user.full_name,
            user_ip_address=current_ip,
            user_system_name=current_system_name,
            image_filename=image_filename
        )
        db.session.add(ticket)
        db.session.commit()

        # Create attachment records for non-image files
        for attachment_filename in other_attachments:
            attachment = Attachment(
                ticket_id=ticket.id,
                filename=attachment_filename
            )
            db.session.add(attachment)
        
        if other_attachments:
            db.session.commit()

        # Send email notification for ticket creation
        from utils.email import send_ticket_creation_notification
        try:
            email_sent = send_ticket_creation_notification(ticket)
            if email_sent:
                flash(f'Ticket {ticket.ticket_number} created successfully and confirmation email sent!', 'success')
            else:
                flash(f'Ticket {ticket.ticket_number} created successfully but confirmation email failed.', 'warning')
        except Exception as e:
            logging.error(f"Error sending ticket creation notification: {e}")
            flash(f'Ticket {ticket.ticket_number} created successfully but confirmation email failed.', 'warning')
        
        return redirect(url_for('user_dashboard'))

    return render_template('create_ticket.html', form=form)

@app.route('/ticket/<int:ticket_id>')
@login_required
def view_ticket(ticket_id):
    """View ticket details"""
    ticket = Ticket.query.get_or_404(ticket_id)
    user = get_current_user()
    
    # Check if user can view this ticket
    can_view = False
    if user.is_super_admin or user.is_admin:
        can_view = True  # Super admins and admins can view all tickets
    elif user.is_hod:
        # HODs can view tickets from their department
        ticket_user = User.query.get(ticket.user_id)
        if ticket_user and ticket_user.department == user.department:
            can_view = True
    elif ticket.user_id == user.id:
        can_view = True  # Users can view their own tickets
    
    if not can_view:
        abort(403)
    
    form = CommentForm()
    assign_form = AssignTicketForm() if user.can_manage_tickets else None
    
    return render_template('view_ticket.html', ticket=ticket, form=form, 
                         assign_form=assign_form, user=user)

@app.route('/ticket/<int:ticket_id>/comment', methods=['POST'])
@login_required
def add_comment(ticket_id):
    """Add comment to ticket"""
    ticket = Ticket.query.get_or_404(ticket_id)
    user = get_current_user()
    
    # Check if user can comment on this ticket
    can_comment = False
    if user.is_super_admin or user.is_admin:
        can_comment = True  # Super admins and admins can comment on all tickets
    elif user.is_hod:
        # HODs can comment on tickets from their department
        ticket_user = User.query.get(ticket.user_id)
        if ticket_user and ticket_user.department == user.department:
            can_comment = True
    elif ticket.user_id == user.id:
        can_comment = True  # Users can comment on their own tickets
    
    if not can_comment:
        abort(403)
    
    form = CommentForm()
    if form.validate_on_submit():
        comment = TicketComment(
            ticket_id=ticket_id,
            user_id=user.id,
            comment=form.comment.data
        )
        db.session.add(comment)
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Send email notification for comment addition
        from utils.email import send_ticket_comment_notification
        try:
            email_sent = send_ticket_comment_notification(ticket, form.comment.data, user)
            if email_sent:
                flash('Comment added successfully and notifications sent!', 'success')
            else:
                flash('Comment added successfully but some notifications failed.', 'warning')
        except Exception as e:
            logging.error(f"Error sending ticket comment notification: {e}")
            flash('Comment added successfully but notifications failed.', 'warning')
    
    return redirect(url_for('view_ticket', ticket_id=ticket_id))

@app.route('/ticket/<int:ticket_id>/edit', methods=['GET', 'POST'])
@super_admin_required
def edit_ticket(ticket_id):
    """Edit ticket (admin only)"""
    ticket = Ticket.query.get_or_404(ticket_id)
    form = UpdateTicketForm(obj=ticket)
    current_user = get_current_user()
    
    if form.validate_on_submit():
        # Only status can be updated - no one can edit title, description, category, or priority
        old_status = ticket.status
        ticket.status = form.status.data
        
        # Set resolved_at if status changed to Resolved
        if old_status != 'Resolved' and ticket.status == 'Resolved':
            ticket.resolved_at = datetime.utcnow()
        elif ticket.status != 'Resolved':
            ticket.resolved_at = None
        
        ticket.updated_at = datetime.utcnow()
        
        # Add comment if super admin provided one
        admin_comment = request.form.get('admin_comment', '').strip()
        if current_user and current_user.is_super_admin and admin_comment:
            comment = TicketComment(
                ticket_id=ticket.id,
                user_id=current_user.id,
                comment=f"Status updated to '{ticket.status}'. {admin_comment}"
            )
            db.session.add(comment)
        elif old_status != ticket.status:
            # Add automatic status change comment
            comment = TicketComment(
                ticket_id=ticket.id,
                user_id=current_user.id,
                comment=f"Status updated from '{old_status}' to '{ticket.status}'"
            )
            db.session.add(comment)
        
        db.session.commit()
        
        # Send email notification for status update
        from utils.email import send_ticket_status_update_notification
        try:
            email_sent = send_ticket_status_update_notification(ticket, old_status, current_user)
            if email_sent:
                flash('Ticket status updated successfully and notifications sent!', 'success')
            else:
                flash('Ticket status updated successfully but some notifications failed.', 'warning')
        except Exception as e:
            logging.error(f"Error sending ticket status update notification: {e}")
            flash('Ticket status updated successfully but notifications failed.', 'warning')
        
        return redirect(url_for('view_ticket', ticket_id=ticket_id))
    
    return render_template('edit_ticket.html', form=form, ticket=ticket, user=current_user)


@app.route('/ticket/<int:ticket_id>/assign', methods=['POST'])
@super_admin_required
def assign_ticket(ticket_id):
    """Assign ticket to admin"""
    ticket = Ticket.query.get_or_404(ticket_id)
    form = AssignTicketForm()

    if form.validate_on_submit():
        current_user = get_current_user()
        ticket.assigned_to = form.assigned_to.data
        ticket.assigned_by = current_user.id if current_user else None
        if ticket.status == 'Open':
            ticket.status = 'In Progress'
        ticket.updated_at = datetime.utcnow()
        ticket.assigned_at = datetime.utcnow()
        db.session.commit()

        assignee = User.query.get(form.assigned_to.data)

        # Send comprehensive email notifications
        from utils.email import send_ticket_assignment_notification
        try:
            email_sent = send_ticket_assignment_notification(ticket, assignee, current_user)
            if email_sent:
                flash(f'Ticket assigned to {assignee.full_name} and notifications sent!', 'success')
            else:
                flash(f'Ticket assigned to {assignee.full_name} but some notifications failed.', 'warning')
        except Exception as e:
            logging.error(f"Error sending assignment notifications: {e}")
            flash(f'Ticket assigned to {assignee.full_name} but notifications failed.', 'warning')

    return redirect(url_for('view_ticket', ticket_id=ticket_id))


@app.route('/edit-user/<int:user_id>', methods=['GET', 'POST'])
@super_admin_required
def edit_user(user_id):
    """Edit user (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)
    form = UserProfileForm()

    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.email = form.email.data
        user.department = form.department.data
        user.specialization = form.specialization.data if form.specialization.data else None
        user.system_name = form.system_name.data
        # Only update password if a new value is provided
        if form.password.data:
            user.password = generate_password_hash(form.password.data)
        db.session.commit()
        flash(f'User {user.username} updated successfully!', 'success')
        return redirect(url_for('view_user', user_id=user_id))

    # Pre-populate the form
    form.username.data = user.username
    form.role.data = user.role
    form.first_name.data = user.first_name
    form.last_name.data = user.last_name
    form.email.data = user.email
    form.department.data = user.department
    form.specialization.data = user.specialization
    form.system_name.data = user.system_name
    # Do not pre-fill password for security reasons

    return render_template('edit_user.html', form=form, user=user)

@app.route('/manage-users')
@super_admin_required
def manage_users():
    """Super Admin user management"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    users = User.query.all()
    return render_template('manage_users.html', users=users)

@app.route('/create-user', methods=['GET', 'POST'])
@super_admin_required
def create_user():
    """Create new user (Super Admin only)"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    form = UserRegistrationForm()
    if form.validate_on_submit():
        new_user = User(
            username=form.username.data,
            email=form.email.data,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            department=form.department.data,
            specialization=form.specialization.data if form.specialization.data else None,
            role=form.role.data
        )
        new_user.set_password(form.password.data)
        db.session.add(new_user)
        db.session.commit()
        
        flash(f'User {new_user.username} created successfully!', 'success')
        return redirect(url_for('manage_users'))
    
    return render_template('create_user.html', form=form)

@app.route('/view-user/<int:user_id>')
@super_admin_required
def view_user(user_id):
    """View user details (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    # Get user's tickets
    user_tickets = Ticket.query.filter_by(user_id=user_id).order_by(Ticket.created_at.desc()).all()
    assigned_tickets = Ticket.query.filter_by(assigned_to=user_id).order_by(Ticket.created_at.desc()).all()
    
    return render_template('view_user.html', user=user, user_tickets=user_tickets, assigned_tickets=assigned_tickets)



@app.route('/delete-user/<int:user_id>', methods=['POST'])
@super_admin_required
def delete_user(user_id):
    """Delete user (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    user_to_delete = User.query.get_or_404(user_id)
    
    # Prevent self-deletion for safety
    if user_to_delete.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('manage_users'))
    
    try:
        # Get user's tickets and reassign or handle them
        user_tickets = Ticket.query.filter_by(user_id=user_id).all()
        assigned_tickets = Ticket.query.filter_by(assigned_to=user_id).all()
        
        # Update tickets created by this user to preserve data integrity
        for ticket in user_tickets:
            ticket.user_name = f"{user_to_delete.full_name} (Deleted User)"
            ticket.user_id = None  # Set to None to indicate deleted user
        
        # Reassign tickets that were assigned to this user
        for ticket in assigned_tickets:
            ticket.assigned_to = None
            ticket.status = 'Open'  # Reset status to Open for reassignment
        
        # Delete user's comments (cascade should handle this, but being explicit)
        comments = TicketComment.query.filter_by(user_id=user_id).all()
        for comment in comments:
            db.session.delete(comment)
        
        # Delete the user
        username = user_to_delete.username
        db.session.delete(user_to_delete)
        db.session.commit()
        
        flash(f'User "{username}" has been successfully deleted. Their tickets have been preserved and reassigned tickets are now available for assignment.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/assign-work/<int:ticket_id>', methods=['GET', 'POST'])
@super_admin_required
def assign_work(ticket_id):
    """Super Admin assigns work to specific admins based on category"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Get all super admins for assignment (simplified role structure)
    admins = User.query.filter_by(role='super_admin').all()
    
    form = AssignTicketForm()
    form.assigned_to.choices = [(admin.id, f"{admin.full_name} ({admin.department})") for admin in admins]
    
    if form.validate_on_submit():
        ticket.assigned_to = form.assigned_to.data
        ticket.assigned_by = user.id
        ticket.status = 'In Progress'
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        assignee = User.query.get(form.assigned_to.data)
        
        # Send comprehensive email notifications for assignment
        from utils.email import send_ticket_assignment_notification
        try:
            email_sent = send_ticket_assignment_notification(ticket, assignee, user)
            if email_sent:
                flash(f'Work assigned to {assignee.full_name} and email notifications sent to all parties!', 'success')
            else:
                flash(f'Work assigned to {assignee.full_name} but some email notifications failed.', 'warning')
        except Exception as e:
            logging.error(f"Error sending assignment email: {e}")
            flash(f'Work assigned to {assignee.full_name} but email notifications failed.', 'warning')
        
        return redirect(url_for('super_admin_dashboard'))
    
    return render_template('assign_work.html', form=form, ticket=ticket, admins=admins)

def create_default_admin():
    """Create default super admin and test user if none exists"""
    try:
        super_admin = User.query.filter_by(role='super_admin').first()
        if not super_admin:
            # Create Super Admin
            super_admin_user = User(
                username='superadmin',
                email='superadmin@gtnengineering.com',
                first_name='Super',
                last_name='Administrator',
                department='IT',
                role='super_admin'
            )
            super_admin_user.set_password('super123')
            db.session.add(super_admin_user)
            db.session.commit()
            
            # Create a test user
            test_user = User(
                username='testuser',
                email='user@gtnengineering.com',
                first_name='Test',
                last_name='User',
                department='Engineering',
                role='user'
            )
            test_user.set_password('test123')
            db.session.add(test_user)
            db.session.commit()
            
            logging.info("Default super admin and test user created")
    except Exception as e:
        logging.error(f"Error creating default users: {e}")
        db.session.rollback()

@app.route('/reports-dashboard')
@super_admin_required
def reports_dashboard():
    """Reports Dashboard with visual analytics (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    # Get comprehensive statistics
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='Open').count()
    in_progress_tickets = Ticket.query.filter_by(status='In Progress').count()
    resolved_tickets = Ticket.query.filter_by(status='Resolved').count()
    closed_tickets = Ticket.query.filter_by(status='Closed').count()
    
    # Category breakdown
    hardware_tickets = Ticket.query.filter_by(category='Hardware').count()
    software_tickets = Ticket.query.filter_by(category='Software').count()
    network_tickets = Ticket.query.filter_by(category='Network').count()
    other_tickets = Ticket.query.filter_by(category='Other').count()
    
    # Priority breakdown  
    critical_tickets = Ticket.query.filter_by(priority='Critical').count()
    high_tickets = Ticket.query.filter_by(priority='High').count()
    medium_tickets = Ticket.query.filter_by(priority='Medium').count()
    low_tickets = Ticket.query.filter_by(priority='Low').count()
    
    # Get all tickets for detailed table
    all_tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()
    
    stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'closed_tickets': closed_tickets,
        'hardware_tickets': hardware_tickets,
        'software_tickets': software_tickets,
        'network_tickets': network_tickets,
        'other_tickets': other_tickets,
        'critical_tickets': critical_tickets,
        'high_tickets': high_tickets,
        'medium_tickets': medium_tickets,
        'low_tickets': low_tickets
    }
    
    # Prepare chart data for JavaScript
    chart_data = {
        'category': [hardware_tickets, software_tickets, network_tickets, other_tickets],
        'priority': [critical_tickets, high_tickets, medium_tickets, low_tickets],
        'status': [open_tickets, in_progress_tickets, resolved_tickets, closed_tickets]
    }
    
    # Get recent tickets for activity timeline
    recent_tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(10).all()
    
    # Get top users by ticket count
    from sqlalchemy import func
    top_users_query = db.session.query(
        User.id,
        User.username,
        User.first_name,
        User.last_name,
        User.department,
        func.count(Ticket.id).label('ticket_count')
    ).join(Ticket, User.id == Ticket.user_id).group_by(User.id, User.username, User.first_name, User.last_name, User.department).order_by(func.count(Ticket.id).desc()).limit(8).all()
    
    # Convert to list of dictionaries for easier template access
    top_users = []
    for user_data in top_users_query:
        top_users.append({
            'id': user_data[0],
            'username': user_data[1],
            'first_name': user_data[2],
            'last_name': user_data[3],
            'department': user_data[4],
            'ticket_count': user_data[5]
        })
    
    return render_template('reports_dashboard.html', 
                         stats=stats, 
                         tickets=all_tickets, 
                         chart_data=chart_data,
                         recent_tickets=recent_tickets,
                         top_users=top_users,
                         total_tickets=total_tickets,
                         open_tickets=open_tickets,
                         in_progress_tickets=in_progress_tickets,
                         resolved_tickets=resolved_tickets,
                         closed_tickets=closed_tickets,
                         hardware_tickets=hardware_tickets,
                         software_tickets=software_tickets,
                         critical_tickets=critical_tickets,
                         high_tickets=high_tickets,
                         medium_tickets=medium_tickets,
                         low_tickets=low_tickets)

@app.route('/edit-assignment/<int:ticket_id>', methods=['GET', 'POST'])
@super_admin_required
def edit_assignment(ticket_id):
    """Edit ticket assignment (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    ticket = Ticket.query.get_or_404(ticket_id)
    
    if request.method == 'POST':
        assigned_to = request.form.get('assigned_to')
        if assigned_to:
            assigned_to = int(assigned_to) if assigned_to != '0' else None
        else:
            assigned_to = None
            
        ticket.assigned_to = assigned_to
        ticket.assigned_by = current_user.id
        ticket.updated_at = datetime.utcnow()
        
        try:
            db.session.commit()
            assignee_name = User.query.get(assigned_to).full_name if assigned_to else 'Unassigned'
            
            # Send comprehensive email notifications if assigned to someone
            if assigned_to:
                from utils.email import send_ticket_assignment_notification
                assignee = User.query.get(assigned_to)
                try:
                    email_sent = send_ticket_assignment_notification(ticket, assignee, current_user)
                    if email_sent:
                        flash(f'Ticket {ticket.ticket_number} assigned to {assignee_name} and email notifications sent to all parties!', 'success')
                    else:
                        flash(f'Ticket {ticket.ticket_number} assigned to {assignee_name} but some email notifications failed.', 'warning')
                except Exception as e:
                    logging.error(f"Error sending assignment email: {e}")
                    flash(f'Ticket {ticket.ticket_number} assigned to {assignee_name} but email notifications failed.', 'warning')
            else:
                flash(f'Ticket {ticket.ticket_number} has been unassigned.', 'success')
                
            return redirect(url_for('super_admin_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash('Error updating assignment. Please try again.', 'error')
            logging.error(f"Error updating ticket assignment: {e}")
    
    # Get all admin users for assignment dropdown
    admin_users = User.query.filter_by(role="super_admin").all()
    
    return render_template('edit_assignment.html', ticket=ticket, admin_users=admin_users)

@app.route('/view-image/<filename>')
@login_required
def view_image(filename):
    """View uploaded ticket image - admins can view any, users can view their own"""
    current_user = get_current_user()
    
    # Find ticket with this image
    ticket = Ticket.query.filter_by(image_filename=filename).first()
    if not ticket:
        abort(404)
    
    # Check permissions - admins/super admins can view any, HODs can view their department, users only their own tickets
    can_view = False
    if current_user.is_super_admin or current_user.is_admin:
        can_view = True
    elif current_user.is_hod:
        ticket_user = User.query.get(ticket.user_id)
        if ticket_user and ticket_user.department == current_user.department:
            can_view = True
    elif ticket.user_id == current_user.id:
        can_view = True
    
    if not can_view:
        abort(403)
    
    try:
        return send_from_directory('uploads', filename)
    except FileNotFoundError:
        abort(404)

@app.route('/download-attachment/<filename>')
@login_required
def download_attachment(filename):
    """Download file attachment - admins can download any, HODs can download their department, users can download their own"""
    current_user = get_current_user()
    
    # Find the attachment record
    attachment = Attachment.query.filter_by(filename=filename).first()
    if not attachment:
        abort(404)
    
    # Check permissions - admins can download any, HODs can download their department, users only their own tickets
    can_download = False
    if current_user.is_super_admin or current_user.is_admin:
        can_download = True
    else:
        ticket = Ticket.query.get(attachment.ticket_id)
        if ticket:
            if current_user.is_hod:
                ticket_user = User.query.get(ticket.user_id)
                if ticket_user and ticket_user.department == current_user.department:
                    can_download = True
            elif ticket.user_id == current_user.id:
                can_download = True
    
    if not can_download:
        abort(403)
    
    try:
        return send_from_directory('uploads', filename, as_attachment=True)
    except FileNotFoundError:
        abort(404)

@app.route('/download-excel-report')
@super_admin_required
def download_excel_report():
    """Download Excel report of all tickets (Super Admin only) with filtering options"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))

    try:
        # --- FILTER PARAMS ---
        filter_mode = request.args.get('filter_mode', 'range')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        month = request.args.get('month')
        year = request.args.get('year')

        # --- BUILD QUERY BASED ON FILTER ---
        query = Ticket.query.join(User, Ticket.user_id == User.id)

        if filter_mode == 'range' and from_date and to_date:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d')
            to_dt = datetime.strptime(to_date, '%Y-%m-%d')
            query = query.filter(Ticket.created_at >= from_dt, Ticket.created_at <= to_dt)
        elif filter_mode == 'month' and month:
            y, m = map(int, month.split('-'))
            query = query.filter(
                extract('year', Ticket.created_at) == y,
                extract('month', Ticket.created_at) == m
            )
        elif filter_mode == 'year' and year:
            query = query.filter(extract('year', Ticket.created_at) == int(year))

        tickets = query.all()

        # --- EXCEL GENERATION USING GTN ENGINEERING TEMPLATE ---
        # Load the GTN Engineering template
        template_path = 'attached_assets/R05 - USER COMPLAINT REGISTER_1752222026172.xlsx'
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        ws.title = "R05 C"
        
        # Preserve the template structure - update the date field
        current_date = datetime.now().strftime('%d.%m.%Y')
        ws.cell(row=2, column=21, value=current_date)  # Update date in the template
        
        # Template headers are in row 3 - mapping to helpdesk data:
        # SNo, Dept, Issue, Issue Type, (empty cols), Raised By, Originating Date, Mode of communication, (empty), 
        # Issue Cleared By, Issue clearance date, Reason of the issue, (empty cols), Solution / Action taken, (empty cols), Final Remarks
        
        # Add ticket data starting from row 4 (template rows 4+ are for data)
        data_start_row = 4
        for idx, ticket in enumerate(tickets, 1):
            row = data_start_row + idx - 1
            
            # Map helpdesk data to GTN template columns
            assignee_name = ticket.assignee.full_name if hasattr(ticket, 'assignee') and ticket.assignee else 'Unassigned'
            user_name = getattr(ticket, 'user_name', ticket.user.full_name if hasattr(ticket.user, 'full_name') else 'N/A')
            user_dept = getattr(ticket.user, 'department', 'N/A') or 'N/A'
            
            # Fill the template columns with ticket data
            ws.cell(row=row, column=1, value=idx)  # SNo
            ws.cell(row=row, column=2, value=user_dept)  # Dept
            ws.cell(row=row, column=3, value=ticket.title)  # Issue
            ws.cell(row=row, column=4, value=ticket.category)  # Issue Type
            ws.cell(row=row, column=8, value=user_name)  # Raised By
            ws.cell(row=row, column=9, value=utc_to_ist(ticket.created_at).strftime('%d.%m.%Y') if ticket.created_at else 'N/A')  # Originating Date
            ws.cell(row=row, column=10, value='Online Portal')  # Mode of communication
            ws.cell(row=row, column=12, value=assignee_name)  # Issue Cleared By
            ws.cell(row=row, column=13, value=utc_to_ist(ticket.resolved_at).strftime('%d.%m.%Y') if ticket.resolved_at else 'Pending')  # Issue clearance date
            ws.cell(row=row, column=14, value=ticket.description)  # Reason of the issue
            ws.cell(row=row, column=18, value=f"Status: {ticket.status}")  # Solution / Action taken
            ws.cell(row=row, column=21, value=f"Priority: {ticket.priority}")  # Final Remarks
        
        # Auto-adjust column widths for better readability - handle merged cells
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row in range(1, ws.max_row + 1):
                try:
                    cell = ws.cell(row=row, column=col_idx)
                    if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp matching GTN template format
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'R05 - USER COMPLAINT REGISTER_{timestamp}.xlsx'

        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        return response

    except Exception as e:
        logging.error(f"Error generating Excel report: {e}")
        flash('Error generating report. Please try again.', 'error')
        return redirect(url_for('reports_dashboard'))

# Master Data Management Routes
@app.route('/super_admin/master_data')
@super_admin_required
def master_data_dashboard():
    """Master Data management dashboard"""
    categories = MasterDataCategory.query.all()
    priorities = MasterDataPriority.query.order_by(MasterDataPriority.level).all()
    statuses = MasterDataStatus.query.all()
    email_settings = EmailSettings.query.first()
    timezone_settings = TimezoneSettings.query.first()
    backup_settings = BackupSettings.query.first()
    users = User.query.order_by(User.created_at.desc()).all()
    users_count = User.query.count()
    
    # Email notification statistics
    from models import EmailNotificationLog
    email_sent_count = EmailNotificationLog.query.filter_by(status='sent').count()
    email_failed_count = EmailNotificationLog.query.filter_by(status='failed').count()
    
    return render_template('master_data/dashboard.html',
                         categories=categories,
                         priorities=priorities, 
                         statuses=statuses,
                         email_settings=email_settings,
                         timezone_settings=timezone_settings,
                         backup_settings=backup_settings,
                         users=users,
                         users_count=users_count,
                         email_sent_count=email_sent_count,
                         email_failed_count=email_failed_count)


@app.route('/super_admin/master_data/categories', methods=['GET', 'POST'])
@super_admin_required
def manage_categories():
    """Manage ticket categories"""
    form = MasterDataCategoryForm()
    categories = MasterDataCategory.query.all()
    
    if form.validate_on_submit():
        category = MasterDataCategory(
            name=form.name.data,
            description=form.description.data,
            is_active=form.is_active.data
        )
        db.session.add(category)
        db.session.commit()
        flash(f'Category "{category.name}" created successfully!', 'success')
        return redirect(url_for('manage_categories'))
    
    return render_template('master_data/categories.html', form=form, categories=categories)


@app.route('/super_admin/master_data/categories/<int:category_id>/edit', methods=['GET', 'POST'])
@super_admin_required
def edit_category(category_id):
    """Edit a category"""
    category = MasterDataCategory.query.get_or_404(category_id)
    form = MasterDataCategoryForm(obj=category)
    
    if form.validate_on_submit():
        category.name = form.name.data
        category.description = form.description.data
        category.is_active = form.is_active.data
        category.updated_at = datetime.utcnow()
        db.session.commit()
        flash(f'Category "{category.name}" updated successfully!', 'success')
        return redirect(url_for('manage_categories'))
    
    return render_template('master_data/edit_category.html', form=form, category=category)


@app.route('/super_admin/master_data/categories/<int:category_id>/delete', methods=['POST'])
@super_admin_required
def delete_category(category_id):
    """Delete a category"""
    category = MasterDataCategory.query.get_or_404(category_id)
    category_name = category.name
    
    try:
        db.session.delete(category)
        db.session.commit()
        flash(f'Category "{category_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting category: {str(e)}', 'error')
    
    return redirect(url_for('manage_categories'))


@app.route('/super_admin/master_data/priorities', methods=['GET', 'POST'])
@super_admin_required
def manage_priorities():
    """Manage ticket priorities"""
    form = MasterDataPriorityForm()
    priorities = MasterDataPriority.query.order_by(MasterDataPriority.level).all()
    
    if form.validate_on_submit():
        priority = MasterDataPriority(
            name=form.name.data,
            description=form.description.data,
            level=form.level.data,
            color_code=form.color_code.data,
            is_active=form.is_active.data
        )
        db.session.add(priority)
        db.session.commit()
        flash(f'Priority "{priority.name}" created successfully!', 'success')
        return redirect(url_for('manage_priorities'))
    
    return render_template('master_data/priorities.html', form=form, priorities=priorities)


@app.route('/super_admin/master_data/statuses', methods=['GET', 'POST'])
@super_admin_required
def manage_statuses():
    """Manage ticket statuses"""
    form = MasterDataStatusForm()
    statuses = MasterDataStatus.query.all()
    
    if form.validate_on_submit():
        status = MasterDataStatus(
            name=form.name.data,
            description=form.description.data,
            color_code=form.color_code.data,
            is_active=form.is_active.data
        )
        db.session.add(status)
        db.session.commit()
        flash(f'Status "{status.name}" created successfully!', 'success')
        return redirect(url_for('manage_statuses'))
    
    return render_template('master_data/statuses.html', form=form, statuses=statuses)


@app.route('/super_admin/master_data/priorities/<int:priority_id>/edit', methods=['GET', 'POST'])
@super_admin_required
def edit_priority(priority_id):
    """Edit a priority"""
    priority = MasterDataPriority.query.get_or_404(priority_id)
    form = MasterDataPriorityForm(obj=priority)
    
    if form.validate_on_submit():
        priority.name = form.name.data
        priority.description = form.description.data
        priority.level = form.level.data
        priority.color_code = form.color_code.data
        priority.is_active = form.is_active.data
        priority.updated_at = datetime.utcnow()
        db.session.commit()
        flash(f'Priority "{priority.name}" updated successfully!', 'success')
        return redirect(url_for('manage_priorities'))
    
    return render_template('master_data/edit_priority.html', form=form, priority=priority)


@app.route('/super_admin/master_data/priorities/<int:priority_id>/delete', methods=['POST'])
@super_admin_required
def delete_priority(priority_id):
    """Delete a priority"""
    priority = MasterDataPriority.query.get_or_404(priority_id)
    priority_name = priority.name
    
    try:
        db.session.delete(priority)
        db.session.commit()
        flash(f'Priority "{priority_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting priority: {str(e)}', 'error')
    
    return redirect(url_for('manage_priorities'))


@app.route('/super_admin/master_data/statuses/<int:status_id>/edit', methods=['GET', 'POST'])
@super_admin_required
def edit_status(status_id):
    """Edit a status"""
    status = MasterDataStatus.query.get_or_404(status_id)
    form = MasterDataStatusForm(obj=status)
    
    if form.validate_on_submit():
        status.name = form.name.data
        status.description = form.description.data
        status.color_code = form.color_code.data
        status.is_active = form.is_active.data
        status.updated_at = datetime.utcnow()
        db.session.commit()
        flash(f'Status "{status.name}" updated successfully!', 'success')
        return redirect(url_for('manage_statuses'))
    
    return render_template('master_data/edit_status.html', form=form, status=status)


@app.route('/super_admin/master_data/statuses/<int:status_id>/delete', methods=['POST'])
@super_admin_required
def delete_status(status_id):
    """Delete a status"""
    status = MasterDataStatus.query.get_or_404(status_id)
    status_name = status.name
    
    try:
        db.session.delete(status)
        db.session.commit()
        flash(f'Status "{status_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting status: {str(e)}', 'error')
    
    return redirect(url_for('manage_statuses'))





@app.route('/super_admin/master_data/email_settings', methods=['GET', 'POST'])
@super_admin_required
def manage_email_settings():
    """Manage SMTP email settings"""
    form = EmailSettingsForm()
    email_settings = EmailSettings.query.first()
    
    if form.validate_on_submit():
        if email_settings:
            # Update existing settings
            email_settings.smtp_server = form.smtp_server.data
            email_settings.smtp_port = form.smtp_port.data
            email_settings.smtp_username = form.smtp_username.data
            email_settings.smtp_password = form.smtp_password.data
            email_settings.use_tls = form.use_tls.data
            email_settings.from_email = form.from_email.data
            email_settings.from_name = form.from_name.data
            email_settings.is_active = form.is_active.data
            email_settings.updated_at = datetime.utcnow()
        else:
            # Create new settings
            email_settings = EmailSettings(
                smtp_server=form.smtp_server.data,
                smtp_port=form.smtp_port.data,
                smtp_username=form.smtp_username.data,
                smtp_password=form.smtp_password.data,
                use_tls=form.use_tls.data,
                from_email=form.from_email.data,
                from_name=form.from_name.data,
                is_active=form.is_active.data
            )
            db.session.add(email_settings)
        
        db.session.commit()
        flash('Email settings saved successfully!', 'success')
        return redirect(url_for('manage_email_settings'))
    elif email_settings:
        # Pre-populate form with existing data
        form.smtp_server.data = email_settings.smtp_server
        form.smtp_port.data = email_settings.smtp_port
        form.smtp_username.data = email_settings.smtp_username
        form.smtp_password.data = email_settings.smtp_password
        form.use_tls.data = email_settings.use_tls
        form.from_email.data = email_settings.from_email
        form.from_name.data = email_settings.from_name
        form.is_active.data = email_settings.is_active
    
    return render_template('master_data/email_settings.html', form=form, email_settings=email_settings)

@app.route('/super_admin/test_email_settings', methods=['GET', 'POST'])
@login_required
@super_admin_required
def test_email_settings():
    """Test email configuration by sending a test email"""
    try:
        current_user = get_current_user()
        from utils.email import send_assignment_email
        
        # Send test email to current user
        email_sent = send_assignment_email(
            current_user.email, 
            "TEST", 
            current_user.full_name
        )
        
        if email_sent:
            flash('Test email sent successfully! Check your inbox.', 'success')
        else:
            flash('Test email failed to send. Check email settings and application logs.', 'error')
            
    except Exception as e:
        logging.error(f"Test email error: {e}")
        flash('Test email failed with error. Check application logs for details.', 'error')
        
    return redirect(url_for('manage_email_settings'))

@app.route('/super_admin/master_data/timezone_settings', methods=['GET', 'POST'])
@super_admin_required
def manage_timezone_settings():
    """Manage timezone settings"""
    form = TimezoneSettingsForm()
    timezone_settings = TimezoneSettings.query.first()
    
    if form.validate_on_submit():
        if timezone_settings:
            # Update existing settings
            timezone_settings.timezone_name = form.timezone_name.data
            timezone_settings.display_name = form.display_name.data
            timezone_settings.utc_offset = form.utc_offset.data
            timezone_settings.is_active = form.is_active.data
            timezone_settings.updated_at = datetime.utcnow()
        else:
            # Create new settings
            timezone_settings = TimezoneSettings(
                timezone_name=form.timezone_name.data,
                display_name=form.display_name.data,
                utc_offset=form.utc_offset.data,
                is_active=form.is_active.data
            )
            db.session.add(timezone_settings)
        
        db.session.commit()
        flash('Timezone settings saved successfully!', 'success')
        return redirect(url_for('manage_timezone_settings'))
    elif timezone_settings:
        # Pre-populate form with existing data
        form.timezone_name.data = timezone_settings.timezone_name
        form.display_name.data = timezone_settings.display_name
        form.utc_offset.data = timezone_settings.utc_offset
        form.is_active.data = timezone_settings.is_active
    
    return render_template('master_data/timezone_settings.html', form=form, timezone_settings=timezone_settings)

@app.route('/super_admin/master_data/backup_settings', methods=['GET', 'POST'])
@super_admin_required
def manage_backup_settings():
    """Manage backup settings"""
    form = BackupSettingsForm()
    backup_settings = BackupSettings.query.first()
    
    if form.validate_on_submit():
        if backup_settings:
            # Update existing settings
            backup_settings.backup_frequency = form.backup_frequency.data
            backup_settings.backup_time = form.backup_time.data
            backup_settings.backup_location = form.backup_location.data
            backup_settings.max_backups = form.max_backups.data
            backup_settings.compress_backups = form.compress_backups.data
            backup_settings.include_attachments = form.include_attachments.data
            backup_settings.email_notifications = form.email_notifications.data
            backup_settings.notification_email = form.notification_email.data
            backup_settings.is_active = form.is_active.data
            backup_settings.updated_at = datetime.utcnow()
        else:
            # Create new settings
            backup_settings = BackupSettings(
                backup_frequency=form.backup_frequency.data,
                backup_time=form.backup_time.data,
                backup_location=form.backup_location.data,
                max_backups=form.max_backups.data,
                compress_backups=form.compress_backups.data,
                include_attachments=form.include_attachments.data,
                email_notifications=form.email_notifications.data,
                notification_email=form.notification_email.data,
                is_active=form.is_active.data
            )
            db.session.add(backup_settings)
        
        db.session.commit()
        flash('Backup settings saved successfully!', 'success')
        return redirect(url_for('manage_backup_settings'))
    elif backup_settings:
        # Pre-populate form with existing data
        form.backup_frequency.data = backup_settings.backup_frequency
        form.backup_time.data = backup_settings.backup_time
        form.backup_location.data = backup_settings.backup_location
        form.max_backups.data = backup_settings.max_backups
        form.compress_backups.data = backup_settings.compress_backups
        form.include_attachments.data = backup_settings.include_attachments
        form.email_notifications.data = backup_settings.email_notifications
        form.notification_email.data = backup_settings.notification_email
        form.is_active.data = backup_settings.is_active
    
    return render_template('master_data/backup_settings.html', form=form, backup_settings=backup_settings)

@app.route('/super_admin/master_data/email_notifications')
@super_admin_required
def email_notifications_dashboard():
    """Email notifications dashboard to track sent/failed emails"""
    # Get statistics
    total_notifications = EmailNotificationLog.query.count()
    sent_notifications = EmailNotificationLog.query.filter_by(status='sent').count()
    failed_notifications = EmailNotificationLog.query.filter_by(status='failed').count()
    
    # Get recent notifications (last 50)
    recent_notifications = EmailNotificationLog.query.order_by(EmailNotificationLog.created_at.desc()).limit(50).all()
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    message_type_filter = request.args.get('message_type', 'all')
    
    # Build filtered query
    query = EmailNotificationLog.query
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    if message_type_filter != 'all':
        query = query.filter_by(message_type=message_type_filter)
    
    filtered_notifications = query.order_by(EmailNotificationLog.created_at.desc()).limit(100).all()
    
    stats = {
        'total_notifications': total_notifications,
        'sent_notifications': sent_notifications,
        'failed_notifications': failed_notifications,
        'success_rate': round((sent_notifications / total_notifications * 100) if total_notifications > 0 else 0, 1)
    }
    
    return render_template('master_data/email_notifications.html', 
                         stats=stats, 
                         notifications=filtered_notifications,
                         status_filter=status_filter,
                         message_type_filter=message_type_filter)


# Initialize default admin on first import
with app.app_context():
    create_default_admin()

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('403.html'), 403

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500
